# -*- coding: utf-8 -*-
"""macro_adapter.py
   Lee la hoja 'Macro' (.xlsx/.xlsm) y produce:
   (df_factura, df_conceptos, df_forma_pago, df_conceptos_texto)
   [CORREGIDO V7 - ¡LA BUENA!] Mantiene lógica original 100% + cierre seguro con finally wb.close().
"""
import os, re, unicodedata, numpy as np, pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
# Importar Workbook para type hinting (opcional pero bueno)
from openpyxl.workbook import Workbook


def _norm_invoice_id(x: object) -> str:
    s = str(x).strip()
    if re.fullmatch(r"\d+(?:\.0+)?", s):
        try:
            return str(int(float(s)))
        except Exception:
            return s
    return s

PREFERRED_SHEETS = ["Macro", "MACRO", "Hoja1", "Resumen"]

EXCEL_COLS = {
    "num_factura": "A", "fecha_emision": "B", "cif_emisor": "E",
    "cliente_nombre": "G", "cliente_nif": "H", "cliente_dir": "I",
    "cliente_cp_prov": "J",
    "desc_1": "K", "imp_1": "L",
    "desc_2": "M", "imp_2": "N",
    "desc_3": "O", "imp_3": "P",
    "desc_4": "Q", "imp_4": "R",
    "desc_5": "S", "imp_5": "T",
    "desc_6": "U", "imp_6": "V",
    "desc_7": "W", "imp_7": "X",
    "desc_8": "Y", "imp_8": "Z",
    "suplidos_aa": "AA", "iban_macro": "AB", "estado": "AC",
    "base_ad": "AD", "total_ah": "AH", "factura_original": "AI",
}

ISO2_TO_ISO3 = {
    "AT":"AUT","BE":"BEL","BG":"BGR","CY":"CYP","CZ":"CZE","DE":"DEU","DK":"DNK",
    "EE":"EST","ES":"ESP","FI":"FIN","FR":"FRA","GR":"GRC","EL":"GRC","HR":"HRV",
    "HU":"HUN","IE":"IRL","IT":"ITA","LT":"LTU","LU":"LUX","LV":"LVA","MT":"MLT",
    "NL":"NLD","PL":"POL","PT":"PRT","RO":"ROU","SE":"SWE","SI":"SVN","SK":"SVK",
    "GB":"GBR","UK":"GBR","NO":"NOR","CH":"CHE","US":"USA","CN":"CHN"
}

def excel_col_to_idx(col_letters: str) -> int:
    # Usa la función original de openpyxl que sí existe
    return column_index_from_string(col_letters.strip().upper()) - 1

def clean_nif_cliente(nif: str) -> str:
    if nif is None: return ""
    s = str(nif)
    s = re.sub(r'^\s*(CIF|NIF)\s*[:\-]?\s*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', '', s)
    return s

def normalize_cif_emisor(s: str) -> str:
    if s is None: return ""
    s = str(s).upper().strip()
    s = re.sub(r'^(CIF|NIF)\s*', '', s)
    s = re.sub(r'^ES', '', s)
    s = re.sub(r'[\s\-\._]', '', s)
    return s

def normalize_series_list(series_str: str) -> list:
    if pd.isna(series_str): return []
    return [p.strip() for p in str(series_str).split(',') if p.strip() != ""]

def coerce_number(x):
    if x is None: return 0.0
    if isinstance(x, (int, float, np.number)):
        return float(x) if not np.isnan(x) else 0.0
    s = str(x).strip()
    if s == "": return 0.0
    # Permitir notación científica y controlar errores
    if re.match(r"^-?\d+(\.\d+)?([eE][+-]?\d+)?$", s):
        try:
            val = float(s)
            return val if not np.isnan(val) else 0.0
        except ValueError:
            pass  # Continuar para probar con reemplazo de comas

    # Lógica original de reemplazo de comas (ajustada)
    s_cleaned = s.replace('.', '').replace(',', '.')
    try:
        val = float(s_cleaned)
        return val if not np.isnan(val) else 0.0
    except ValueError:
        return 0.0

def _normalize_simple_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("�", "n")
    return text

def _unique_headers(header_row):
    seen = {}
    result = []
    for idx, raw in enumerate(header_row):
        name = str(raw or f"col_{idx}").strip()
        if not name:
            name = f"col_{idx}"
        base = name
        count = seen.get(base, 0)
        if count:
            name = f"{base}_{count+1}"
        seen[base] = count + 1
        result.append(name)
    return result

def _split_cp_prov(s: str):
    """Acepta '41004 Sevilla' o 'Sevilla 41004'."""
    s = str(s or "").strip()
    if not s: return "", ""
    t = s.split()
    if t and re.fullmatch(r'\d{4,6}', t[-1]):  # ... CP
        return t[-1], " ".join(t[:-1]).strip()
    if t and re.fullmatch(r'\d{4,6}', t[0]):   # CP ...
        return t[0], " ".join(t[1:]).strip()
    return "", s

# [MODIFICADO] Usar try...finally para asegurar wb.close()
def _read_sheet_to_df_any(path: str, preferred_names=None):
    wb: Workbook = None # Definir wb fuera del try
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        names = [ws.title for ws in wb.worksheets]
        target = None
        if preferred_names:
            low = {n.lower(): n for n in names}
            for cand in preferred_names:
                if cand.lower() in low:
                    target = low[cand.lower()]; break
        if not target:
            for ws in wb.worksheets:
                if ws.sheet_state == "visible": target = ws.title; break
            if not target: target = wb.worksheets[0].title
        ws = wb[target]
        rows_raw = [list(r) for r in ws.iter_rows(values_only=True)]
        
        # Convertir objetos datetime a strings de forma segura para evitar problemas de timestamp
        from datetime import datetime as py_datetime
        rows = []
        for row in rows_raw:
            converted_row = []
            for cell_val in row:
                # Si es un objeto datetime de Python, convertir a string de forma segura
                if isinstance(cell_val, py_datetime):
                    try:
                        # NUNCA llamar a strftime() directamente - puede causar overflow
                        # Usar solo la representación string del objeto
                        date_repr = str(cell_val)
                        # Si tiene formato de fecha (YYYY-MM-DD HH:MM:SS o YYYY-MM-DD), extraerlo
                        if len(date_repr) >= 10 and date_repr[4] == '-' and date_repr[7] == '-':
                            # Extraer solo la parte de fecha (primeros 10 caracteres)
                            date_str = date_repr[:10]
                            # Validar el año desde el string
                            year = int(date_str[:4])
                            if 1900 <= year <= 2100:
                                converted_row.append(date_str)
                            else:
                                # Año fuera de rango, usar None
                                converted_row.append(None)
                        else:
                            # No tiene formato de fecha reconocible, usar None
                            converted_row.append(None)
                    except (OverflowError, OSError, ValueError, AttributeError, TypeError):
                        # Si hay cualquier error, usar None
                        converted_row.append(None)
                else:
                    converted_row.append(cell_val)
            rows.append(converted_row)
    except Exception as e:
        # Imprimir error pero también propagarlo
        print(f"Error leyendo hoja genérica de {path}: {e}")
        raise # Es importante relanzar el error para que main.py lo capture
    finally:
        # ---> AÑADIDO <---
        if wb: # Solo intentar cerrar si el workbook se llegó a abrir
            try:
                wb.close() # <- Cierra el archivo
            except Exception as close_err:
                 print(f"Advertencia: Error al intentar cerrar {path} en _read_sheet_to_df_any: {close_err}")
        # ---> FIN AÑADIDO <---

    # --- El archivo debería estar cerrado aquí ---
    # Lógica original para crear DataFrame (sin tocar)
    if not rows: return pd.DataFrame()
    max_len = max(len(r) for r in rows) if rows else 0
    norm_rows = [r + [None]*(max_len-len(r)) for r in rows]
    return pd.DataFrame(norm_rows)


# [MODIFICADO] Usar try...finally para asegurar wb.close()
def _read_clientes_df_from_same_book(macro_path: str, sheet_name_candidates=None) -> pd.DataFrame:
    """Lee la hoja CLIENTES (o variantes) del MISMO archivo Excel que Macro (no crea columnas nuevas)."""
    if sheet_name_candidates is None:
        sheet_name_candidates = ["CLIENTES", "Clientes", "clientes", "EMISORES", "EMISOR", "CONFIG", "Config"]
    wb: Workbook = None # Definir wb fuera del try
    try:
        wb = load_workbook(macro_path, data_only=True, read_only=True)
        names = [ws.title for ws in wb.worksheets]
        low = {n.lower(): n for n in names}
        target = None
        for cand in sheet_name_candidates:
            if cand.lower() in low:
                target = low[cand.lower()]
                break
        if not target:
            raise ValueError("No se encontró la hoja 'CLIENTES' en el Excel.")
        ws = wb[target]
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        print(f"Error leyendo hoja 'CLIENTES' de {macro_path}: {e}")
        raise # Propagar error
    finally:
        # ---> AÑADIDO <---
        if wb: # Solo cerrar si se abrió
            try:
                wb.close() # <- Cierra archivo
            except Exception as close_err:
                print(f"Advertencia: Error cerrar {macro_path} en clientes: {close_err}")
        # ---> FIN AÑADIDO <---

    # --- El archivo debería estar cerrado aquí ---
    # Lógica original para crear DataFrame (sin tocar)
    if not rows:
        return pd.DataFrame()
    headers = [str(c or "").strip() for c in rows[0]]
    data = [list(r) for r in rows[1:]]
    df = pd.DataFrame(data, columns=headers)
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df

# [MODIFICADO] Usar try...finally para asegurar wb.close()
def _read_emisores_df(emisores_path: str) -> pd.DataFrame:
    wb: Workbook = None # Definir wb fuera del try
    try:
        wb = load_workbook(emisores_path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        print(f"Error leyendo archivo de emisores {emisores_path}: {e}")
        raise # Propagar error
    finally:
        # ---> AÑADIDO <---
        if wb: # Solo cerrar si se abrió
            try:
                wb.close() # <- Cierra archivo
            except Exception as close_err:
                 print(f"Advertencia: Error al intentar cerrar {emisores_path} en _read_emisores_df: {close_err}")
        # ---> FIN AÑADIDO <---

    # --- El archivo debería estar cerrado aquí ---
    # Lógica original para crear DataFrame (sin tocar)
    if not rows: return pd.DataFrame()
    headers = [str(c or "").strip() for c in rows[0]]
    data = [list(r) for r in rows[1:]]
    df = pd.DataFrame(data, columns=headers)
    df.columns = [str(c).strip().lower() for c in df.columns]
    if "empresa_nombre" not in df.columns and "nombre_legal" in df.columns:
        df["empresa_nombre"] = df["nombre_legal"]
    if "pais_iso2" not in df.columns and "pains_iso2" in df.columns:
        df["pais_iso2"] = df["pains_iso2"]
    if "cif_aliases" not in df.columns:
        df["cif_aliases"] = ""
    return df


# --- El resto de funciones (_match_emisor, _snap_vat, adapt_from_macro) se mantienen EXACTAMENTE IGUAL que tu versión original ---
def _match_emisor(df_emisores: pd.DataFrame, cif_input: str):
    target = normalize_cif_emisor(cif_input)
    # Añadir comprobación por si target queda vacío después de normalizar
    if not target:
        print(f"Advertencia: CIF de entrada '{cif_input}' normalizado a vacío, no se puede buscar.")
        return None, None

    df_emisores = df_emisores.copy()
    # Comprobar si 'cif' existe, si no, intentar con 'cif/nif' o similares
    cif_col_name = None
    possible_cif_cols = ['cif', 'cif/nif', 'nif', 'vat']
    for col in possible_cif_cols:
        if col in df_emisores.columns:
            cif_col_name = col
            break
    if not cif_col_name:
        print(f"Advertencia: La hoja 'CLIENTES' no contiene ninguna columna de CIF esperada ({', '.join(possible_cif_cols)}).")
        return None, None

    df_emisores["_cif_norm"] = df_emisores[cif_col_name].astype(str).map(normalize_cif_emisor)
    hits = df_emisores[df_emisores["_cif_norm"] == target]
    has_token_col = "api_token" in df_emisores.columns # Comprobar si existe la columna token

    if not hits.empty:
        if has_token_col: # Ordenar solo si la columna existe
            # ESTA ES LA LÓGICA ORIGINAL QUE TENÍAS PARA ORDENAR
            hits = hits.assign(_has_token=hits["api_token"].astype(str).str.strip().ne(""))
            hits = hits.sort_values(by=["_has_token"], ascending=True)
            return hits.iloc[-1], "cif" # Devolver el último (prioriza True si ascending=True)
        else:
            return hits.iloc[0], "cif" # Si no hay token, devolver el primero

    # Búsqueda por alias (igual que antes)
    def split_aliases(s):
        if pd.isna(s) or str(s).strip()=="": return []
        return [a.strip() for a in str(s).split(",") if a.strip()!=""]

    alias_rows = []
    if "cif_aliases" in df_emisores.columns: # Comprobar si existe la columna de alias
        for _, row in df_emisores.iterrows():
            # Asegurarse que row['cif_aliases'] no sea None antes de split
            aliases_str = row.get("cif_aliases","")
            if pd.notna(aliases_str):
                 aliases = [normalize_cif_emisor(a) for a in split_aliases(aliases_str)]
                 if target in aliases: alias_rows.append(row)

    if alias_rows:
        hits = pd.DataFrame(alias_rows)
        if has_token_col: # Ordenar si existe token
             # Lógica original de ordenación
             hits = hits.assign(_has_token=lambda d: d["api_token"].astype(str).str.strip().ne(""))
             hits = hits.sort_values(by=["_has_token"], ascending=True)
             return hits.iloc[-1], "alias" # Devolver el último
        else:
             return hits.iloc[0], "alias" # Si no hay token, devolver el primero

    return None, None # No se encontró ni por CIF ni por alias

def _snap_vat(p):
    """Ajusta el IVA a tipos comunes si está muy cerca."""
    if p is None or (isinstance(p, float) and np.isnan(p)): return 0.0
    try: # Añadir try-except para conversión robusta
        p = float(p)
    except (ValueError, TypeError):
        return 0.0

    for c in (0.0, 4.0, 5.0, 10.0, 21.0):
        # Asegurar que p no sea NaN antes de comparar
        if not np.isnan(p) and abs(p - c) <= 0.25:
            return c
    # Redondear solo si p no es NaN
    return round(p, 2) if not np.isnan(p) else 0.0

# --- adapt_from_macro (lógica principal, SIN CAMBIOS respecto a tu versión original) ---
def adapt_from_macro(macro_path: str):
    df_all = _read_sheet_to_df_any(macro_path, preferred_names=PREFERRED_SHEETS)
    if df_all.empty or df_all.shape[0] < 2: # Mantener la comprobación original
        raise ValueError("La hoja 'Macro' está vacía o no tiene datos")

    # Lógica original para obtener 'df' desde 'df_all'
    df = df_all.iloc[1:].copy().reset_index(drop=True)
    # Asignar nombres de columna desde la primera fila de df_all (lógica original implícita)
    df.columns = [str(c or f"col_{i}").strip() for i, c in enumerate(df_all.iloc[0])]

    cols = {}
    for key, letter in EXCEL_COLS.items():
        idx = excel_col_to_idx(letter)
        # Lógica original para seleccionar columnas
        cols[key] = df.iloc[:, idx] if idx < df.shape[1] else pd.Series([np.nan]*len(df))
    m = pd.DataFrame(cols)

    m["num_factura"] = m["num_factura"].map(_norm_invoice_id)
    # Lógica original de filtrado
    m = m[m["num_factura"]!=""].reset_index(drop=True)
    if m.empty: # Añadir comprobación por si el filtrado deja m vacío
         raise ValueError("No se encontraron filas con número de factura válido tras filtrar.")

    m["tipo"] = np.where(m["num_factura"].str.startswith("Int"), "intereses",
                  np.where(m["num_factura"].str.startswith("A"), "intra", "normal"))
    # Convertir fechas con validación para evitar errores de timestamp fuera de rango
    # IMPORTANTE: Validar ANTES de convertir a datetime para evitar overflow
    def _pre_validate_and_convert_date(date_val):
        """Valida y convierte una fecha de forma segura, evitando overflow"""
        if pd.isna(date_val):
            return pd.NaT
        
        # Si ya es un pd.Timestamp, validar sin acceder a atributos que puedan causar overflow
        if isinstance(date_val, pd.Timestamp):
            try:
                # NUNCA acceder a atributos del timestamp - solo usar string
                date_str = str(date_val)
                # Extraer año del string de forma segura
                if len(date_str) >= 4:
                    # Si tiene formato YYYY-MM-DD, extraer directamente
                    if len(date_str) >= 10 and date_str[4] == '-' and date_str[7] == '-':
                        year = int(date_str[:4])
                    else:
                        # Buscar año en el string
                        import re
                        year_match = re.search(r'\b(19\d{2}|20\d{2})\b', date_str)
                        if year_match:
                            year = int(year_match.group(1))
                        else:
                            return pd.NaT
                    
                    if 1900 <= year <= 2100:
                        return date_val
                return pd.NaT
            except (ValueError, OverflowError, OSError, TypeError):
                return pd.NaT
        
        # Si es un objeto datetime de Python, convertir a string primero
        from datetime import datetime as py_datetime
        if isinstance(date_val, py_datetime):
            try:
                # NUNCA acceder a .year directamente - usar strftime
                date_str = date_val.strftime("%Y-%m-%d")
                # Validar el año desde el string
                year = int(date_str[:4])
                if 1900 <= year <= 2100:
                    # Crear pd.Timestamp desde string (más seguro)
                    return pd.Timestamp(date_str)
                else:
                    return pd.NaT
            except (OverflowError, OSError, ValueError, AttributeError):
                # Si falla strftime, intentar desde representación string
                try:
                    date_repr = str(date_val)
                    if len(date_repr) >= 10 and date_repr[4] == '-' and date_repr[7] == '-':
                        year = int(date_repr[:4])
                        if 1900 <= year <= 2100:
                            return pd.Timestamp(date_repr[:10])
                except:
                    pass
                return pd.NaT
        
        # Si es string, validar año antes de convertir
        if isinstance(date_val, str):
            date_val_clean = str(date_val).strip()
            # Buscar año en el string
            import re
            year_match = re.search(r'\b(19\d{2}|20\d{2})\b', date_val_clean)
            if year_match:
                year = int(year_match.group(1))
                if year < 1900 or year > 2100:
                    return pd.NaT
        
        # Si es número (fecha de Excel), validar rango
        if isinstance(date_val, (int, float)):
            # Números de fecha de Excel suelen estar entre 1 y ~100000
            if date_val < 0 or date_val > 100000:
                return pd.NaT
        
        # Ahora intentar convertir con pd.to_datetime, pero con validación previa más estricta
        try:
            # Si es un número muy grande, podría ser un timestamp Unix en milisegundos
            if isinstance(date_val, (int, float)):
                # Si es un número muy grande (> 1e10), podría ser un timestamp en milisegundos
                if date_val > 1e10:
                    # Probablemente es un timestamp en milisegundos, convertir a segundos
                    date_val = date_val / 1000.0
                # Validar que el timestamp esté en rango razonable (1970-2100)
                if date_val > 4102444800:  # 2100-01-01 en timestamp Unix
                    return pd.NaT
            
            # Intentar convertir con pd.to_datetime
            result = pd.to_datetime(date_val, errors="coerce")
            if pd.isna(result):
                return pd.NaT
            
            # Validar resultado sin acceder a atributos que puedan causar overflow
            # NUNCA acceder a .year, .month, etc. - usar solo string
            try:
                result_str = str(result)
                # Extraer año del string (primeros 4 caracteres si es formato YYYY-MM-DD)
                if len(result_str) >= 4:
                    # Intentar extraer año del string
                    if result_str[4] == '-' and len(result_str) >= 10:
                        # Formato YYYY-MM-DD
                        year = int(result_str[:4])
                    else:
                        # Buscar año en el string
                        import re
                        year_match = re.search(r'\b(19\d{2}|20\d{2})\b', result_str)
                        if year_match:
                            year = int(year_match.group(1))
                        else:
                            return pd.NaT
                    
                    if 1900 <= year <= 2100:
                        return result
                return pd.NaT
            except (ValueError, OverflowError, OSError, TypeError):
                return pd.NaT
                
        except (OverflowError, OSError, ValueError) as e:
            # Si hay error de overflow, retornar NaT
            return pd.NaT
        except Exception:
            return pd.NaT
    
    try:
        # Aplicar validación y conversión segura
        m["fecha_emision"] = m["fecha_emision"].apply(_pre_validate_and_convert_date)
    except Exception as e:
        # Si hay error general, marcar todas como NaT
        import logging
        logging.warning(f"Error al convertir fechas: {e}. Marcando como NaT.")
        m["fecha_emision"] = pd.NaT
    m["cliente_nif_limpio"] = m["cliente_nif"].apply(clean_nif_cliente)
    for c in [f"imp_{i}" for i in range(1,9)] + ["suplidos_aa","base_ad","total_ah"]:
        # Asegurar que la columna exista antes de aplicar coerce_number
        if c in m.columns:
            m[c] = m[c].apply(coerce_number)
        else:
            m[c] = np.nan # Si no existe, llenarla con NaN

    # Emisor desde hoja CLIENTES del mismo Excel
    try:
        df_emisores = _read_clientes_df_from_same_book(macro_path)
        if df_emisores.empty:
            raise ValueError("La hoja CLIENTES está vacía o mal formada")
    except ValueError as e:
         raise ValueError(f"Error al leer la hoja 'CLIENTES': {e}")

    # --- INICIO REFACTORIZACIÓN PARA MÚLTIPLES EMISORES ---

    # 1. Normalizar CIFs en el dataframe principal 'm' para poder agrupar
    def _norm_cif_cell(x):
        if pd.isna(x): return ""
        s = str(x).strip()
        if s == "" or s.lower() in ("none", "nan", "null", "#n/a", "#n/d", "-", "—"):
            return ""
        return normalize_cif_emisor(s)

    if 'cif_emisor' not in m.columns:
         raise ValueError("La columna 'cif_emisor' (E) no se encontró en la hoja Macro procesada.")
    m['cif_emisor_norm'] = m['cif_emisor'].apply(_norm_cif_cell)

    # Listas para acumular los dataframes de cada emisor
    all_facturas = []
    all_conceptos = []
    all_formas_pago = []
    all_textos = []

    # 2. Iterar sobre cada grupo de facturas (agrupadas por CIF de emisor)
    for cif_norm, m_group in m.groupby('cif_emisor_norm'):
        if not cif_norm:
            continue # Omitir filas sin CIF de emisor

        # --- Lógica de procesamiento existente, ahora aplicada a 'm_group' ---
        emisor_row_series, match_type = _match_emisor(df_emisores, cif_norm)
        if emisor_row_series is None:
            # En lugar de lanzar un error, podríamos registrarlo y continuar
            print(f"Advertencia: Empresa no configurada en hoja CLIENTES para CIF: {cif_norm}. Omitiendo {len(m_group)} facturas.")
            continue
        emisor_row = emisor_row_series.to_dict()

        empresa_nombre = str(emisor_row.get("empresa_nombre","") or "").strip()
        if not empresa_nombre:
            empresa_nombre = str(emisor_row.get("nombre_legal","") or "").strip()
        if not empresa_nombre:
             empresa_nombre = cif_norm
             print(f"Advertencia: No se encontró nombre para emisor {cif_norm}, usando CIF.")

        unidad_def   = str(emisor_row.get("unidad_medida_defecto","") or "ud")
        bic_conf     = str(emisor_row.get("bic","") or "").replace(" ","")
        series_list  = normalize_series_list(emisor_row.get("series_retencion",""))
        iban_defecto = str(emisor_row.get("iban_defecto","") or "").replace(" ","")

        def _pick(row_dict, keys):
            for k in keys:
                if k in row_dict:
                    v = str(row_dict.get(k, "") or "").strip()
                    if v: return v
            return ""
        
        # Leer configuración de API: primero del Excel, luego variables de entorno, luego QSettings guardado
        def _get_api_config_from_settings():
            """Obtiene configuración de API desde QSettings como fallback."""
            try:
                import sys
                from PySide6.QtCore import QSettings
                # Usar la misma lógica que AppSettings para encontrar el archivo config.ini
                if getattr(sys, "frozen", False):
                    app_dir = os.path.dirname(sys.executable)
                else:
                    # En desarrollo, buscar el archivo config.ini en la raíz del proyecto
                    current_file = os.path.abspath(__file__)
                    app_dir = os.path.dirname(os.path.dirname(current_file))
                config_path = os.path.join(app_dir, "config.ini")
                settings = QSettings(config_path, QSettings.IniFormat)
                return {
                    "api_token": settings.value("api/token", "").strip() if settings.value("api/token") else "",
                    "api_email": settings.value("api/user", "").strip() if settings.value("api/user") else "",
                    "api_url": settings.value("api/url", "").strip() if settings.value("api/url") else "",
                }
            except Exception:
                return {"api_token": "", "api_email": "", "api_url": ""}
        
        saved_config = _get_api_config_from_settings()
        
        api_token = (_pick(emisor_row, ["api_token","api_key","token","facturantia_token","token_api"]) 
                    or os.environ.get("API_TOKEN", "").strip() 
                    or saved_config.get("api_token", ""))
        api_email = (_pick(emisor_row, ["api_email","email_api","usuario_email","api_user_email","user_email"]) 
                    or os.environ.get("API_EMAIL", "").strip() 
                    or saved_config.get("api_email", ""))
        api_url   = (_pick(emisor_row, ["api_url","url_api","endpoint","api_endpoint"]) 
                    or os.environ.get("API_URL", "").strip() 
                    or saved_config.get("api_url", "")
                    or "https://www.facturantia.com/API/proformas_receptor.php")

        m_group["iban_resuelto"] = m_group["iban_macro"].astype(str).str.replace(" ", "").fillna('')
        mask_empty_iban = m_group["iban_resuelto"].str.strip() == ""
        m_group.loc[mask_empty_iban, "iban_resuelto"] = iban_defecto

        if (m_group["iban_resuelto"].fillna('').str.strip() == "").any():
            rows_missing = m_group[m_group["iban_resuelto"].fillna('').str.strip() == ""]
            idxs = (rows_missing.index + 2).tolist()
            print(f"Advertencia: Falta IBAN para CIF {cif_norm} en filas Excel: {idxs}. Omitiendo estas facturas.")
            continue

        conceptos_rows, textos_rows = [], []
        for _, row in m_group.iterrows():
            num = _norm_invoice_id(row["num_factura"])
            pos = 0
            for i in range(1, 9):
                desc = row.get(f"desc_{i}", "")
                imp  = row.get(f"imp_{i}", np.nan)
                if isinstance(desc, float) and np.isnan(desc): desc = ""
                desc = str(desc or "").strip()
                imp_float = coerce_number(imp)
                is_valid_imp = not pd.isna(imp_float) and imp_float != 0.0
                if desc and is_valid_imp:
                    conceptos_rows.append({
                        "NumFactura": num, "empresa_emisora": empresa_nombre,
                        "descripcion": desc, "cuenta_contable": "7050000",
                        "unidad_medida": unidad_def, "unidades": 1.0,
                        "base_unidad": float(imp_float),
                        "tipo_impuesto": "IVA", "porcentaje": 0.0,
                        "__col_index": i,  # Guardar índice original de columna
                    })
                    pos += 1
                elif desc:
                    textos_rows.append({
                        "NumFactura": num, "empresa_emisora": empresa_nombre,
                        "descripcion": desc, "posicion": pos,
                        "__col_index": i,  # Guardar índice original de columna
                    })
                    pos += 1

        # Identificar la primera columna con descripción (con o sin importe) para cada factura
        df_conceptos_group = pd.DataFrame(conceptos_rows) if conceptos_rows else pd.DataFrame()
        df_txt_group = pd.DataFrame(textos_rows) if textos_rows else pd.DataFrame()
        
        # Combinar todos los índices de columna para encontrar el mínimo por factura
        all_col_indices = []
        if not df_conceptos_group.empty:
            all_col_indices.append(df_conceptos_group[["NumFactura", "empresa_emisora", "__col_index"]])
        if not df_txt_group.empty:
            all_col_indices.append(df_txt_group[["NumFactura", "empresa_emisora", "__col_index"]])
        
        if all_col_indices:
            df_all_cols = pd.concat(all_col_indices, ignore_index=True)
            df_min_col = df_all_cols.groupby(["NumFactura", "empresa_emisora"])["__col_index"].min().reset_index()
            df_min_col.columns = ["NumFactura", "empresa_emisora", "__min_col"]
            
            # Aplicar mayúsculas a conceptos con importe
            if not df_conceptos_group.empty:
                df_conceptos_group = df_conceptos_group.merge(df_min_col, on=["NumFactura", "empresa_emisora"], how="left")
                mask_first = df_conceptos_group["__col_index"] == df_conceptos_group["__min_col"]
                df_conceptos_group.loc[mask_first, "descripcion"] = df_conceptos_group.loc[mask_first, "descripcion"].astype(str).str.upper()
                df_conceptos_group.drop(columns=["__col_index", "__min_col"], inplace=True)
            
            # Aplicar mayúsculas a textos sin importe
            if not df_txt_group.empty:
                df_txt_group = df_txt_group.merge(df_min_col, on=["NumFactura", "empresa_emisora"], how="left")
                mask_first_txt = df_txt_group["__col_index"] == df_txt_group["__min_col"]
                df_txt_group.loc[mask_first_txt, "descripcion"] = df_txt_group.loc[mask_first_txt, "descripcion"].astype(str).str.upper()
                df_txt_group.drop(columns=["__col_index", "__min_col"], inplace=True)

        iva_map = {}
        if not m_group.empty:
            for num, grp in m_group.groupby("num_factura"):
                if grp.empty: continue
                first_row = grp.iloc[0]; tipo = first_row["tipo"]
                aa = float(coerce_number(first_row.get("suplidos_aa", 0.0)) or 0.0)
                ad = float(coerce_number(first_row.get("base_ad", 0.0)) or 0.0)
                ah = float(coerce_number(first_row.get("total_ah", 0.0)) or 0.0)
                if tipo == "normal":
                    vat = None
                    if ad != 0 and ah != 0:
                        try:
                            raw = ((ah - aa) - ad) / ad * 100.0
                            vat = _snap_vat(raw)
                        except ZeroDivisionError: vat = 21.0
                    if vat is None or vat <= 0.0 or np.isnan(vat): vat = 21.0
                    iva_map[num] = float(vat)
                else: iva_map[num] = 0.0
        if not df_conceptos_group.empty:
            df_conceptos_group["porcentaje"] = df_conceptos_group["NumFactura"].map(iva_map).fillna(21.0)

        if not df_conceptos_group.empty:
            df_conceptos_group["tipo_impuesto_retenido"] = ""
            df_conceptos_group["porcentaje_retenido"] = 0.0
            tipo_map = m_group.drop_duplicates(subset=['num_factura']).set_index('num_factura')['tipo'].to_dict()
            def needs_ret(num):
                t = tipo_map.get(num, "normal")
                if t == "intereses": return True
                if t == "normal" and any(str(num).startswith(str(p)) for p in series_list if p): return True
                return False
            mask = df_conceptos_group["NumFactura"].apply(needs_ret)
            df_conceptos_group.loc[mask, "tipo_impuesto_retenido"] = "IRPF"; df_conceptos_group.loc[mask, "porcentaje_retenido"] = 19.0

        fact_rows = []; processed_nums = set()
        for _, row in m_group.iterrows():
            num = _norm_invoice_id(row["num_factura"])
            if num in processed_nums: continue
            cp, prov = _split_cp_prov(row.get("cliente_cp_prov","")); prov = (prov or "").strip()
            nif_clean = str(row["cliente_nif_limpio"] or ""); tipo = str(row["tipo"])
            fecha_emision_val = row["fecha_emision"]
            if pd.isna(fecha_emision_val):
                 print(f"Advertencia: Fecha emisión inválida para factura {num}. Omitida."); processed_nums.add(num); continue
            
            # Obtener año de forma segura sin acceder directamente a .year (puede causar overflow)
            try:
                # Intentar obtener el año desde la representación string
                fecha_str = str(fecha_emision_val)
                if len(fecha_str) >= 4:
                    ejercicio_val = int(fecha_str[:4])
                    # Validar que el año esté en rango razonable
                    if ejercicio_val < 1900 or ejercicio_val > 2100:
                        print(f"Advertencia: Año fuera de rango para factura {num}: {ejercicio_val}. Usando año actual.")
                        from datetime import datetime
                        ejercicio_val = datetime.now().year
                else:
                    # Si no podemos extraer el año del string, usar año actual
                    from datetime import datetime
                    ejercicio_val = datetime.now().year
            except (ValueError, TypeError, AttributeError, OverflowError, OSError) as e:
                # Si falla, usar año actual como fallback
                print(f"Advertencia: Error al obtener año de fecha para factura {num}: {e}. Usando año actual.")
                from datetime import datetime
                ejercicio_val = datetime.now().year
            if tipo == "intra":
                mcc = re.match(r"^([A-Z]{2})", nif_clean); iso2 = mcc.group(1).upper() if mcc else "ES"
                codigo_pais = ISO2_TO_ISO3.get(iso2, "ESP"); tipo_doc = "otro_id"; tipo_res = "U"
            else: codigo_pais = "ESP"; tipo_doc = "nif"; tipo_res = "R"
            pobl_use = prov or ""
            plantilla_emitidas = str(emisor_row.get("plantilla_facturas_emitidas", "") or "").strip()
            plantilla_proforma = str(emisor_row.get("plantilla_facturas_proforma", "") or "").strip()
            fact_rows.append({
                "NumFactura": num,
                "empresa_emisora": empresa_nombre,
                "api_key": api_token,
                "api_email": api_email,
                "api_url": api_url,
                "serie_factura": "",
                "fecha_emision": fecha_emision_val,
                "fecha_vencimiento": fecha_emision_val,
                "descripcion_general": "",
                "tipo_factura": "F1",
                "ejercicio": ejercicio_val,
                "cliente_tipo_persona": "J",
                "cliente_nombre": row["cliente_nombre"],
                "cliente_tipo_documento": tipo_doc,
                "cliente_numero_documento": nif_clean,
                "cliente_cuenta_contable": "4300000",
                "cliente_observacion": "",
                "cliente_tipo_residencia": tipo_res,
                "cliente_codigo_pais": codigo_pais,
                "cliente_provincia": (prov or "")[:20],
                "cliente_poblacion": (pobl_use or "")[:50],
                "cliente_domicilio": row["cliente_dir"],
                "cliente_domicilio_2": "",
                "cliente_cp": cp,
                "cliente_telefono": "",
                "cliente_email": "",
                "total_suplidos": coerce_number(row.get("suplidos_aa", 0.0)) if tipo == "normal" else 0.0,
                "total_gastos_financieros": 0.0,
                "total_retenciones": 0.0,
                "plantilla_facturas_emitidas": plantilla_emitidas,
                "plantilla_facturas_proforma": plantilla_proforma,
                "suplidos_aa": coerce_number(row.get("suplidos_aa", 0.0)),
                "base_ad": coerce_number(row.get("base_ad", 0.0)),
                "total_ah": coerce_number(row.get("total_ah", 0.0)),
            })
            processed_nums.add(num)
        df_factura_group = pd.DataFrame(fact_rows)

        fp_rows = []; facturas_validas = set(df_factura_group["NumFactura"].unique()) if not df_factura_group.empty else set()
        for _, row in m_group.iterrows():
            num = _norm_invoice_id(row["num_factura"])
            if num not in facturas_validas: continue
            iban_final = str(row.get("iban_resuelto","") or "").strip()
            if not iban_final: continue
            fp_rows.append({"NumFactura": num, "empresa_emisora": empresa_nombre, "metodo": "transferencia","transferencia_banco": "ABANCA", "transferencia_beneficiario": empresa_nombre,"transferencia_concepto": "Pago Factura", "transferencia_iban": iban_final,"transferencia_bic": bic_conf if bic_conf else "CAGLESMMXXX"})
        df_forma_pago_group = pd.DataFrame(fp_rows)

        # Append results for this group to the master lists
        all_facturas.append(df_factura_group)
        all_conceptos.append(df_conceptos_group)
        all_formas_pago.append(df_forma_pago_group)
        all_textos.append(df_txt_group)

    # 3. Concatenar los dataframes de todos los emisores
    df_factura = pd.concat(all_facturas, ignore_index=True) if all_facturas else pd.DataFrame()
    if not df_factura.empty:
        df_factura = df_factura.loc[:, ~df_factura.columns.duplicated()]

    df_conceptos = pd.concat(all_conceptos, ignore_index=True) if all_conceptos else pd.DataFrame()
    if not df_conceptos.empty:
        df_conceptos = df_conceptos.loc[:, ~df_conceptos.columns.duplicated()]
        if "__id_norm__" not in df_conceptos.columns and "NumFactura" in df_conceptos.columns:
            df_conceptos["__id_norm__"] = df_conceptos["NumFactura"].map(_norm_invoice_id)

    df_forma_pago = pd.concat(all_formas_pago, ignore_index=True) if all_formas_pago else pd.DataFrame()
    if not df_forma_pago.empty:
        df_forma_pago = df_forma_pago.loc[:, ~df_forma_pago.columns.duplicated()]

    df_txt = pd.concat(all_textos, ignore_index=True) if all_textos else pd.DataFrame()
    if not df_txt.empty:
        df_txt = df_txt.loc[:, ~df_txt.columns.duplicated()]

    # Asegurar que no haya columnas duplicadas antes de reindexar (evita "Reindexing only valid with uniquely valued Index")
    def _dedupe_columns(df: pd.DataFrame) -> pd.DataFrame:
        if df is not None and not df.empty and df.columns.duplicated().any():
            return df.loc[:, ~df.columns.duplicated()]
        return df

    df_factura = _dedupe_columns(df_factura)
    df_conceptos = _dedupe_columns(df_conceptos)
    df_forma_pago = _dedupe_columns(df_forma_pago)
    df_txt = _dedupe_columns(df_txt)

    # --- FIN REFACTORIZACIÓN ---

    # Asegurar columnas esperadas (lógica original)
    expected_fact_cols = ["NumFactura", "empresa_emisora", "api_key", "api_email", "api_url", "serie_factura", "fecha_emision", "fecha_vencimiento", "descripcion_general", "tipo_factura", "ejercicio", "cliente_tipo_persona", "cliente_nombre", "cliente_tipo_documento", "cliente_numero_documento", "cliente_cuenta_contable", "cliente_observacion", "cliente_tipo_residencia", "cliente_codigo_pais", "cliente_provincia", "cliente_poblacion", "cliente_domicilio", "cliente_domicilio_2", "cliente_cp", "cliente_telefono", "cliente_email", "total_suplidos", "total_gastos_financieros", "total_retenciones", "suplidos_aa", "base_ad", "total_ah", "plantilla_facturas_emitidas", "plantilla_facturas_proforma"]
    expected_conc_cols = ["NumFactura", "empresa_emisora", "descripcion", "cuenta_contable", "unidad_medida", "unidades", "base_unidad", "tipo_impuesto", "porcentaje", "tipo_impuesto_retenido", "porcentaje_retenido"]
    expected_fp_cols = ["NumFactura", "empresa_emisora", "metodo", "transferencia_banco", "transferencia_beneficiario", "transferencia_concepto", "transferencia_iban", "transferencia_bic"]
    expected_txt_cols = ["NumFactura", "empresa_emisora", "descripcion", "posicion"]

    def _ensure_columns(df: pd.DataFrame | None, expected_cols: list[str]) -> pd.DataFrame:
        if df is None or df.empty:
            return pd.DataFrame(columns=expected_cols)
        df = df.copy()
        for col in expected_cols:
            if col not in df.columns:
                df[col] = np.nan
        return df[expected_cols]

    df_factura = _ensure_columns(df_factura, expected_fact_cols)
    df_conceptos = _ensure_columns(df_conceptos, expected_conc_cols)
    df_forma_pago = _ensure_columns(df_forma_pago, expected_fp_cols)
    df_txt = _ensure_columns(df_txt, expected_txt_cols)

    # --- [NUEVO] LEER Y PROCESAR HOJAS DE HISTORIAL ---
    # El objetivo es crear df_factura_historico y df_conceptos_historico para que
    # prueba.py pueda buscar facturas originales aunque hayan sido borradas de la hoja "Macro".

    df_factura_historico = pd.DataFrame()
    df_conceptos_historico = pd.DataFrame()

    try:
        wb = load_workbook(macro_path, data_only=True, read_only=True)
        all_sheet_names = wb.sheetnames

        # Excluir las hojas ya procesadas o de configuración
        sheets_to_exclude = PREFERRED_SHEETS + ["CLIENTES", "Clientes", "clientes", "EMISORES", "EMISOR", "CONFIG", "Config"]
        sheets_to_exclude_lower = [s.lower() for s in sheets_to_exclude]

        historical_sheet_names = [
            name for name in all_sheet_names
            if name.lower() not in sheets_to_exclude_lower
        ]

        if historical_sheet_names:
            historical_dfs_raw = []
            for sheet_name in historical_sheet_names:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows or len(rows) < 2:
                    continue
                header_first = rows[0][0] if rows[0] else ""
                normalized_first = _normalize_simple_text(header_first)
                if not normalized_first or not any(token in normalized_first for token in ("factura", "abono")):
                    # Evitamos procesar hojas auxiliares (Tarifas, análisis, etc.)
                    continue
                # Asumimos la misma estructura: cabecera en fila 1, datos desde fila 2
                headers = _unique_headers(rows[0])
                data = rows[1:]
                df_sheet = pd.DataFrame(data, columns=headers)
                historical_dfs_raw.append(df_sheet)

            if historical_dfs_raw:
                # Filtrar columnas completamente vacías o todas-NA antes de concatenar
                # Esto evita el FutureWarning de pandas sobre concatenación con columnas vacías
                historical_dfs_cleaned = []
                for df in historical_dfs_raw:
                    if df.empty:
                        continue
                    # Eliminar columnas que estén completamente vacías o todas-NA
                    df_cleaned = df.dropna(axis=1, how='all')
                    # También eliminar columnas que solo tengan valores vacíos (strings vacíos)
                    # Filtrar columnas que tengan al menos un valor no vacío y no NA
                    non_empty_cols = [
                        col for col in df_cleaned.columns
                        if df_cleaned[col].notna().any() and 
                           (df_cleaned[col].astype(str).str.strip() != '').any()
                    ]
                    if non_empty_cols:
                        df_cleaned = df_cleaned[non_empty_cols]
                    if not df_cleaned.empty:
                        historical_dfs_cleaned.append(df_cleaned)
                
                if historical_dfs_cleaned:
                    df_hist_all = pd.concat(historical_dfs_cleaned, ignore_index=True)
                else:
                    df_hist_all = pd.DataFrame()

                # --- Re-aplicar la misma lógica de procesamiento que para la hoja "Macro" ---
                # Asegurar índice único antes de resetear (evita error "Reindexing only valid with uniquely valued Index")
                df_hist = df_hist_all.copy()
                if df_hist.columns.duplicated().any():
                    df_hist = df_hist.loc[:, ~df_hist.columns.duplicated()]
                # Si el índice no es único, forzar reset con drop=True
                if not df_hist.index.is_unique:
                    df_hist.index = pd.RangeIndex(len(df_hist))
                df_hist = df_hist.reset_index(drop=True)
                # (No necesitamos la primera fila para nombres de columna porque ya los asignamos)

                cols_hist = {}
                for key, letter in EXCEL_COLS.items():
                    idx = excel_col_to_idx(letter)
                    if idx < df_hist.shape[1]:
                        # Asegurarse de que el nombre de columna exista antes de acceder
                        col_name = df_hist.columns[idx]
                        cols_hist[key] = df_hist[col_name]
                    else:
                        cols_hist[key] = pd.Series([np.nan] * len(df_hist))

                m_hist = pd.DataFrame(cols_hist)
                if "fecha_emision" in m_hist.columns:
                    # Convertir fechas con validación para evitar errores de timestamp fuera de rango
                    try:
                        m_hist["fecha_emision"] = pd.to_datetime(m_hist["fecha_emision"], errors="coerce")
                        # Filtrar fechas fuera de rango válido (1900-2100) usando validación segura
                        def _validate_date_safe(date_val):
                            """Valida una fecha de forma segura sin causar overflow"""
                            if pd.isna(date_val):
                                return pd.NaT
                            try:
                                # Convertir a string y extraer año para validar sin comparar timestamps
                                date_str = str(date_val)
                                if len(date_str) >= 4:
                                    year = int(date_str[:4])
                                    if 1900 <= year <= 2100:
                                        return date_val
                                return pd.NaT
                            except (ValueError, OverflowError, OSError, TypeError):
                                return pd.NaT
                        
                        # Aplicar validación de forma segura
                        m_hist["fecha_emision"] = m_hist["fecha_emision"].apply(_validate_date_safe)
                    except (OverflowError, OSError, ValueError) as e:
                        # Si hay error al convertir, marcar todas como NaT
                        import logging
                        logging.warning(f"Error al convertir fechas históricas: {e}. Marcando como NaT.")
                        m_hist["fecha_emision"] = pd.NaT
                for c in [f"imp_{i}" for i in range(1, 9)] + ["suplidos_aa", "base_ad", "total_ah"]:
                    if c in m_hist.columns:
                        m_hist[c] = m_hist[c].apply(coerce_number)
                m_hist["num_factura"] = m_hist["num_factura"].map(_norm_invoice_id)
                m_hist = m_hist[m_hist["num_factura"] != ""]
                # Asegurar índice único antes de resetear
                if not m_hist.empty:
                    if m_hist.columns.duplicated().any():
                        m_hist = m_hist.loc[:, ~m_hist.columns.duplicated()]
                    if not m_hist.index.is_unique:
                        m_hist.index = pd.RangeIndex(len(m_hist))
                    m_hist = m_hist.reset_index(drop=True)

                if not m_hist.empty:
                    m_hist['cif_emisor_norm'] = m_hist['cif_emisor'].apply(_norm_cif_cell)

                    hist_all_facturas = []
                    hist_all_conceptos = []

                    for cif_norm, m_hist_group in m_hist.groupby('cif_emisor_norm'):
                        if not cif_norm: continue

                        emisor_row_series, _ = _match_emisor(df_emisores, cif_norm)
                        if emisor_row_series is None: continue

                        emisor_row = emisor_row_series.to_dict()
                        empresa_nombre = str(emisor_row.get("empresa_nombre","") or cif_norm).strip()
                        unidad_def = str(emisor_row.get("unidad_medida_defecto","") or "ud")

                        # Procesar conceptos del historial
                        conceptos_rows_hist = []
                        textos_rows_hist = []
                        iva_map_hist = {}
                        if not m_hist_group.empty:
                            for num_hist, grp_hist in m_hist_group.groupby("num_factura"):
                                first_hist = grp_hist.iloc[0]
                                base_hist = coerce_number(first_hist.get("base_ad", 0.0))
                                total_hist = coerce_number(first_hist.get("total_ah", 0.0))
                                suplidos_hist = coerce_number(first_hist.get("suplidos_aa", 0.0))
                                iva_hist = total_hist - suplidos_hist - base_hist
                                vat_hist = round((iva_hist / base_hist) * 100.0, 2) if base_hist else 0.0
                                iva_map_hist[_norm_invoice_id(num_hist)] = vat_hist

                        for _, row in m_hist_group.iterrows():
                            num = _norm_invoice_id(row["num_factura"])
                            for i in range(1, 9):
                                desc = row.get(f"desc_{i}", "")
                                imp  = row.get(f"imp_{i}", np.nan)
                                if isinstance(desc, float) and np.isnan(desc): desc = ""
                                desc = str(desc or "").strip()
                                imp_float = coerce_number(imp)
                                is_valid_imp = not pd.isna(imp_float) and imp_float != 0.0
                                if desc and is_valid_imp:
                                    conceptos_rows_hist.append({
                                        "NumFactura": num, "empresa_emisora": empresa_nombre,
                                        "base_unidad": float(imp_float),
                                        "descripcion": str(desc), "unidad_medida": unidad_def,
                                        "__col_index": i,  # Guardar índice original de columna
                                    })
                                elif desc:
                                    textos_rows_hist.append({
                                        "NumFactura": num, "empresa_emisora": empresa_nombre,
                                        "descripcion": str(desc),
                                        "__col_index": i,  # Guardar índice original de columna
                                    })

                        # Identificar la primera columna con descripción (con o sin importe) para cada factura histórica
                        df_conc_hist = pd.DataFrame(conceptos_rows_hist) if conceptos_rows_hist else pd.DataFrame()
                        df_txt_hist = pd.DataFrame(textos_rows_hist) if textos_rows_hist else pd.DataFrame()
                        
                        # Combinar todos los índices de columna para encontrar el mínimo por factura
                        all_col_indices_hist = []
                        if not df_conc_hist.empty:
                            all_col_indices_hist.append(df_conc_hist[["NumFactura", "empresa_emisora", "__col_index"]])
                        if not df_txt_hist.empty:
                            all_col_indices_hist.append(df_txt_hist[["NumFactura", "empresa_emisora", "__col_index"]])
                        
                        if all_col_indices_hist:
                            df_all_cols_hist = pd.concat(all_col_indices_hist, ignore_index=True)
                            df_min_col_hist = df_all_cols_hist.groupby(["NumFactura", "empresa_emisora"])["__col_index"].min().reset_index()
                            df_min_col_hist.columns = ["NumFactura", "empresa_emisora", "__min_col"]
                            
                            if not df_conc_hist.empty:
                                df_conc_hist["tipo_impuesto"] = "IVA"
                                df_conc_hist["porcentaje"] = df_conc_hist["NumFactura"].map(iva_map_hist).fillna(0.0)
                                df_conc_hist["tipo_impuesto_retenido"] = ""
                                df_conc_hist["porcentaje_retenido"] = 0.0
                                # Aplicar mayúsculas a conceptos con importe
                                df_conc_hist = df_conc_hist.merge(df_min_col_hist, on=["NumFactura", "empresa_emisora"], how="left")
                                mask_first_hist = df_conc_hist["__col_index"] == df_conc_hist["__min_col"]
                                df_conc_hist.loc[mask_first_hist, "descripcion"] = df_conc_hist.loc[mask_first_hist, "descripcion"].astype(str).str.upper()
                                df_conc_hist.drop(columns=["__col_index", "__min_col"], inplace=True)
                            
                            # Aplicar mayúsculas a textos sin importe (aunque no se usen después, se procesan para consistencia)
                            if not df_txt_hist.empty:
                                df_txt_hist = df_txt_hist.merge(df_min_col_hist, on=["NumFactura", "empresa_emisora"], how="left")
                                mask_first_txt_hist = df_txt_hist["__col_index"] == df_txt_hist["__min_col"]
                                df_txt_hist.loc[mask_first_txt_hist, "descripcion"] = df_txt_hist.loc[mask_first_txt_hist, "descripcion"].astype(str).str.upper()
                                df_txt_hist.drop(columns=["__col_index", "__min_col"], inplace=True)
                        elif not df_conc_hist.empty:
                            # Si no hay textos, aplicar lógica simple solo a conceptos
                            df_conc_hist["tipo_impuesto"] = "IVA"
                            df_conc_hist["porcentaje"] = df_conc_hist["NumFactura"].map(iva_map_hist).fillna(0.0)
                            df_conc_hist["tipo_impuesto_retenido"] = ""
                            df_conc_hist["porcentaje_retenido"] = 0.0
                            df_conc_hist["__min_col"] = df_conc_hist.groupby(["NumFactura", "empresa_emisora"])["__col_index"].transform("min")
                            mask_first_hist = df_conc_hist["__col_index"] == df_conc_hist["__min_col"]
                            df_conc_hist.loc[mask_first_hist, "descripcion"] = df_conc_hist.loc[mask_first_hist, "descripcion"].astype(str).str.upper()
                            df_conc_hist.drop(columns=["__col_index", "__min_col"], inplace=True)
                        
                        if not df_conc_hist.empty:
                            hist_all_conceptos.append(df_conc_hist)

                        # Procesar facturas del historial
                        fact_rows_hist = []
                        processed_nums_hist = set()
                        for _, row in m_hist_group.iterrows():
                            num = _norm_invoice_id(row["num_factura"])
                            if num in processed_nums_hist: continue

                            fact_rows_hist.append({
                                "NumFactura": num,
                                "empresa_emisora": empresa_nombre,
                                "cliente_nif": str(row.get("cliente_nif","")),
                                "cliente_nombre": str(row.get("cliente_nombre","")),
                                "cliente_numero_documento": clean_nif_cliente(row.get("cliente_nif","")),
                                "suplidos_aa": coerce_number(row.get("suplidos_aa", 0.0)),
                                "base_ad": coerce_number(row.get("base_ad", 0.0)),
                                "total_ah": coerce_number(row.get("total_ah", 0.0)),
                            })
                            processed_nums_hist.add(num)

                        if fact_rows_hist:
                            hist_all_facturas.append(pd.DataFrame(fact_rows_hist))

                    if hist_all_facturas:
                        df_factura_historico = pd.concat(hist_all_facturas, ignore_index=True)
                    if hist_all_conceptos:
                        df_conceptos_historico = pd.concat(hist_all_conceptos, ignore_index=True)

    except Exception as e:
        # Si falla la lectura del historial, no detenemos el proceso, solo lo advertimos.
        print(f"Advertencia: No se pudo procesar el historial de facturas. Causa: {e}")
    finally:
        if 'wb' in locals() and wb:
            wb.close()

    # Devolver los 4 dataframes originales + los 2 del historial
    return df_factura, df_conceptos, df_forma_pago, df_txt, df_factura_historico, df_conceptos_historico